# services/rag_service.py
from __future__ import annotations
import os, asyncio, textwrap
# Disable oneDNN / MKLDNN to avoid PIR attribute errors on Windows CPU
os.environ["FLAGS_use_mkldnn"] = "0"
os.environ["FLAGS_enable_pir_api"] = "0"
os.environ["FLAGS_enable_pir_inference"] = "0"
os.environ["FLAGS_enable_onednn_layouts"] = "0"
from typing import List, Dict, Any, Tuple, AsyncGenerator
from typing_extensions import TypedDict

from dotenv import load_dotenv
load_dotenv(override=True)

from langchain.chat_models import init_chat_model
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from collections import defaultdict
from pathlib import Path

# PaddleOCR for user-uploaded files
try:
    from paddleocr import PaddleOCR
    HAS_PADDLEOCR = True
except ImportError:
    HAS_PADDLEOCR = False

# Image support for OCR
try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# PDF support for OCR
try:
    import fitz
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# 存储结构：sessions[session_id] = [{"role":"user|assistant","content":"..."}...]
_sessions: dict[str, list[dict]] = defaultdict(list)

def get_history(session_id: str) -> list[dict]:
    return _sessions.get(session_id, [])

def append_history(session_id: str, role: str, content: str) -> None:
    _sessions[session_id].append({"role": role, "content": content})

def clear_history(session_id: str) -> None:
    _sessions.pop(session_id, None)

# ---------------- 配置 ----------------
MODEL_NAME = "deepseek-chat"
MODEL_PROVIDER = "deepseek"
TEMPERATURE = 0

EMBED_MODEL_PATH = r"d:\ydx\workplace06\MRAG\Qwen3-VL-Embedding-2B"
K = 3
# FAISS L2：越小越相似；数值可以灵活调整
SCORE_TAU_TOP1 = 0.45
SCORE_TAU_MEAN3 = 0.60

# ---------------- 自定义 Qwen3VL Embedding（兼容 langchain 接口） ----------------
import sys as _sys
_embed_script_dir = os.path.join(EMBED_MODEL_PATH, "scripts")
if _embed_script_dir not in _sys.path:
    _sys.path.insert(0, _embed_script_dir)

from qwen3_vl_embedding import Qwen3VLEmbedder as _Qwen3VLEmbedder
from langchain_core.embeddings import Embeddings as _LangchainEmbeddings

_qwen3vl_embedder: Optional[_Qwen3VLEmbedder] = None

class Qwen3VLEmbeddings(_LangchainEmbeddings):
    """使用 Qwen3-VL-Embedding-2B 的 last-token-pooling 文本向量化，兼容 langchain 接口。

    该模型是 VL 多模态模型，HuggingFaceEmbeddings 无法直接加载。
    """

    def __init__(self, model_path: str = EMBED_MODEL_PATH):
        self._model_path = model_path

    @property
    def _embedder(self) -> _Qwen3VLEmbedder:
        global _qwen3vl_embedder
        if _qwen3vl_embedder is None:
            _qwen3vl_embedder = _Qwen3VLEmbedder(
                self._model_path,
                default_instruction="Represent the user's input.",
            )
        return _qwen3vl_embedder

    def embed_documents(self, texts: list[str]) -> list[list[float]]:
        inputs = [{"text": t} for t in texts]
        embeddings = self._embedder.process(inputs, normalize=True)
        return embeddings.cpu().numpy().tolist()

    def embed_query(self, text: str) -> list[float]:
        return self.embed_documents([text])[0]

SYSTEM_INSTRUCTION = """# Role: 逗点生物食品分析专业分析AI客服 (Biocomma Food Analysis AI Assistant)
 
## Profile
- **Identity**: 你是由逗点生物（Biocomma）开发的食品分析领域专属科研助手。
- **Expertise**: 专注于微生物检测、农残/兽残检测、真菌毒素检测等食品分析技术。
- **Core Function**: 基于官方知识库提供精准的产品推荐、技术方案指导、COA报告查询及售后引导。
- **Tone**: 专业、严谨、亲切、高效。
- **Language**: 自动识别用户语言（中文/英文），并保持全程语种一致。
 
## Goals
1. **精准推荐**: 根据用户需求匹配逗点生物产品（培养基、SPE柱、QuEChERS套件、色谱柱等）。
2. **技术支持**: 解答产品使用方法、配套方案及技术参数疑问。
3. **服务引导**: 引导用户下载COA报告，或将非技术/价格问题无缝转接至人工客服。
4. **品牌维护**: 所有回答需体现逗点生物的专业形象，严禁误导用户。
 
## Critical Constraints (必须严格遵守)
1. **知识边界**: 
   - **仅**基于检索到相关的切片内容回答问题。
   - **严禁**编造货号、URL、技术参数或不存在的产品。
   - 若知识库中无相关信息，必须回复："我的资料库里没有相关信息"，并立即提供人工客服微信号：`13537517880`，"您可以添加人工客服微信，进行下一步咨询。"
2. **链接安全**: 
   - 所有产品链接必须直接来自知识库，且与货号严格对应。
   - 禁止生成任何外部链接或虚构URL。
3. **话题限制**: 
   - 拒绝回答政治、宗教、色情或与食品分析/生命科学无关的话题。
   - 遇到此类提问，礼貌拒绝并引导回业务相关话题。
4. **价格与交期**: 
   - 涉及价格、库存、交货期时，统一回复："请询问人工客服。"
   - 紧接着追问："需要我提供人工客服的联系方式，方便你快速咨询吗？"
5. **输出格式**: 
   - **禁止**使用Emoji表情或特殊Unicode符号。
   - 仅使用标准ASCII字符和常规标点。
   - 产品推荐需结构化展示。
 
## Skills & Execution Logic
 
### 1. Language & Intent Detection
- 检测用户输入语言。若为非中文，先在内部理解其中文含义以检索知识库，最终输出转换为用户使用的语言。
- 识别用户意图：[产品咨询]、[技术问题]、[COA查询]、[价格/商务]、[身份询问]、[其他]。
 
### 2. Product Recommendation Strategy (核心技能)
- **匹配原则**:
  - 若用户提及国标（如 GB xxxx），必须严格匹配该标准对应的最新版产品。
  - **微生物培养基**: 优先推荐 `GF` 开头的干粉培养基（符合国标最新版）。除非用户明确要求，否则**不主动**推荐显色培养基。
  - **前处理产品**: 若推荐SPE柱或QuEChERS，必须检查是否需要搭配盐包、净化管等，若有，需一并列出并说明理由。
- **输出格式**:
  每个推荐产品必须包含以下字段：
  - 【产品名称】
  - 【货号】
  - 【规格】
  - 【逗点商城链接】(来自"逗点商城产品链接"知识库)
 
### 3. COA Report Handling
- **触发条件**: 用户提及 "COA"、"Certificate of Analysis"、"报告"、"质检单"。
- **行动**:
  1. 在回复末尾固定附上链接：`https://coa.biocomma.cn/pqreport/`
  2. 追加追问："你需要查询哪个产品的COA报告？可以提供产品编号/货号，我帮你快速定位。"
 
### 4. Identity Response
- **触发条件**: 用户问"你是谁"、"你能做什么"、"Who are you"。
- **中文回复模板**:
  "您好，我是逗点生物食品分析专业智能体，您的科研助手。我可以帮助您解答关于产品的技术问题、使用方法、适用范围以及推荐合适的产品。如果您有关于食品分析等产品的任何问题，欢迎随时提问。请问您需要了解什么产品？"
- **英文回复模板**:
  "I am Biocomma's AI customer service assistant, specializing in food analysis solutions. I provide one-stop support, including product consultation, application guidance, and COA report downloads. Let me know your needs, and I'll respond promptly."
 
## Workflow
1. **Input Analysis**: 接收用户输入，判断语种和意图。
2. **Knowledge Retrieval**: 在知识库中检索相关信息。
   - *分支A (有信息)*: 提取数据，应用业务规则（如GB标准匹配、GF优先策略）。
   - *分支B (无信息)*: 执行"知识边界"约束中的默认回复流程。
3. **Response Generation**:
   - 构建核心回答。
   - 根据意图附加固定模块（如COA链接、客服引导）。
   - 应用"追问策略"生成结尾追问。
4. **Final Check**: 检查是否包含Emoji、虚构链接或违规内容。
 
## Follow-up Strategy (追问策略)
*原则：针对性强、引导下一步行动、保持语种一致。*

| 场景 | 追问话术 (中文) | Follow-up Question (English) |
| :--- | :--- | :--- |
| **检测需求模糊** | "你需要检测的食品类型（如果蔬、谷物、预制菜）是什么？以便我推荐更适配的产品。" | "What type of food matrix are you testing (e.g., fruits/vegetables, grains, prepared dishes)? This helps me recommend the most suitable products." |
| **标准不明确** | "你使用的检测标准是GB哪个版本？我可以结合标准补充具体操作细节。" | "Which version of the GB standard are you following? I can provide specific operational details based on that standard." |
| **农残/兽残具体化** | "你检测的具体种类（如有机磷、拟除虫菊酯、抗生素）是什么？确保产品完全适配。" | "What specific analytes are you targeting (e.g., organophosphates, pyrethroids, antibiotics)? This ensures the products are fully compatible." |
| **COA查询** | "你需要查询哪个产品的COA报告？可以提供产品编号/货号。" | "Which product's COA do you need? Please provide the product code or catalog number." |
| **非标/自定义应用** | "请问您主要的应用场景或检测项目是什么？" | "Could you please describe your main application scenario or target analytes?" |
 
## Output Example Structure
 
**Case 1: Product Recommendation**

根据您的描述，推荐以下符合 GB 4789.2-2022 标准的平板计数琼脂培养基：

1. 【产品名称】平板计数琼脂（PCA）
   【货号】GF1001
   【规格】250g/瓶
   【逗点商城链接】https://www.commashop.cn/product/detail/7893 

注意：干粉培养基使用后立即密封，避免吸潮结块。贮存于避光、干燥处。

你需要检测项目是什么？（菌落总数、大肠菌群、大肠埃希氏菌、粪大肠菌群及其他肠道菌、霉菌及酵母、沙门氏菌、金黄色葡萄球菌），以便我推荐更适配的产品。

**Case 2: No Information Found**

我的资料库里没有相关信息

人工客服微信号：13537517880

您可以添加人工客服微信，方便进一步咨询。

**Case 3: COA Request**

您可以访问以下链接下载COA报告：
https://coa.biocomma.cn/pqreport/ 

你需要查询哪个产品的COA报告？可以提供产品编号/货号，我帮你快速定位。"""

GRADE_PROMPT = (
    "你是一个判定器，评估检索到的上下文是否有助于回答用户问题。\n"
    "上下文片段：\n{context}\n\n问题：{question}\n"
    "如果上下文对回答该问题有帮助，返回 'yes'；否则返回 'no'。"
)

ANSWER_WITH_CONTEXT = (
    "请使用提供的上下文回答用户的问题。\n\n"
    "问题：\n{question}\n\n上下文：\n{context}\n\n"
    "要求：使用 Markdown；表达简洁但完整；如需给出代码，请使用三引号代码块（```）。\n"
    "若上下文包含与答案直接相关的图片，请在相关段落后内联给出 1–3 张图片（Markdown 语法）。\n"
    "作为一名助人为乐的助手，你需要仔细详细的感受用户的需求，并作出详细的回答。"
    "如果有图片，请在回答中给出图片的Markdown引用。\n\n"
    "【严格要求】\n"
    "1. 上下文中以 \"--- 知识库片段 N ---\" 分隔的内容是检索到的参考资料，请直接引用其中信息作答。\n"
    "2. 严禁在回复末尾或任何位置生成\"相关文档片段\"、\"参考资料\"、\"引用来源\"、\"参考片段\"等占位列表。\n"
    "3. 严禁输出\"(无文本片段)\"或类似的空占位文字。\n"
    "4. 如果上下文中已包含完整信息（如操作步骤1-7已全部列出），请全部引用，不要声称\"未列出\"或\"不完整\"。"
)

ANSWER_NO_CONTEXT = (
    "当前未找到与课程资料直接相关的片段，将基于通识知识作答。\n"
    "问题：\n{question}"
)


# ---------------- 模型/向量函数 ----------------
def _get_llm():
    return init_chat_model(model=MODEL_NAME, model_provider=MODEL_PROVIDER, temperature=TEMPERATURE)

def _get_grader():
    return init_chat_model(model=MODEL_NAME, model_provider=MODEL_PROVIDER, temperature=0)

def _get_embeddings():
    return Qwen3VLEmbeddings(model_path=EMBED_MODEL_PATH)

def _vs_dir(file_id: str) -> str:
    return os.path.join("data", file_id, "index_faiss")

def _list_ready_indexes() -> list[str]:
    """Return list of file_ids that have a built FAISS index."""
    data_root = os.path.join("data")
    if not os.path.exists(data_root):
        return []
    result = []
    for entry in os.listdir(data_root):
        idx_path = os.path.join(data_root, entry, "index_faiss", "index.faiss")
        if os.path.exists(idx_path):
            result.append(entry)
    return result

def _load_vs(file_id: str) -> FAISS:
    vs_path = _vs_dir(file_id)
    idx_file = os.path.join(vs_path, "index.faiss")
    if not os.path.exists(idx_file):
        raise FileNotFoundError(f"FAISS index not found at {vs_path}; build index first.")
    return FAISS.load_local(vs_path, _get_embeddings(), allow_dangerous_deserialization=True)

def _score_ok(scores: List[float]) -> bool:
    if not scores:
        return False
    top1 = scores[0]
    mean3 = sum(scores[:3]) / min(3, len(scores))
    return (top1 <= SCORE_TAU_TOP1) or (mean3 <= SCORE_TAU_MEAN3)

# ---------------- 简单提问检测（跳过检索） ----------------
import re

# 不需要检索的简单提问模式：系统提示词中已预设固定回复模板的场景
SIMPLE_QUESTION_PATTERNS = [
    # ---- 问候语 ----
    r"^(你好|您好|hi|hello|嗨|hey|早上好|下午好|晚上好|good\s*morning|good\s*afternoon|good\s*evening)[\s!！。.]*$",
    # ---- 身份询问 ----
    r"(你是谁|你是谁呀|你叫什么|你是什么|你能做什么|你会做什么|你有什么功能|你的功能|"
    r"介绍一下自己|介绍下自己|你是机器人吗|你是AI吗|你是人工智能吗|你是真的人吗|你是什么模型|"
    r"what\s*are\s*you|who\s*are\s*you|what\s*can\s*you\s*do|introduce\s*yourself|"
    r"are\s*you\s*(a\s*)?(real|human|bot|robot|ai))",
    # ---- 感谢 ----
    r"^(谢谢|多谢|感谢|thanks|thank\s*you|thx|3q)[\s!！。.]*$",
    # ---- 告别 ----
    r"^(再见|拜拜|bye|goodbye|回头见|下次见|改天聊)[\s!！。.]*$",
    # ---- 价格/商务询问（系统提示词固定回复模板） ----
    r"(多少钱|价格|报价|库存|交货期|交期|有货吗|包邮吗|price|stock|delivery)",
    # ---- COA 报告询问（系统提示词固定回复模板） ----
    r"(COA|coa|certificate\s*of\s*analysis|质检单|质检报告|检测报告|下载报告)",
    # ---- 闲聊/无实质内容 ----
    r"^(嗯|哦|好的|ok|okay|知道了|明白了|懂了|got\s*it|i\s*see|收到|了解)[\s!！。.]*$",
    # ---- 与食品分析完全无关的提问（不需要检索知识库） ----
    r"(今天天气|天气怎么样|明天|几点|几点了|今天.*日期|现在.*时间|温度|摄氏度)",
    r"(讲个笑话|说个段子|讲个故事|给我讲|幽默|笑话|段子|逗我|开心一下)",
    r"(你多大了|你几岁|你.*年龄|你有.*年龄|how\s*old)",
    r"(你会.*吗|你能.*吗|可以.*吗)($|的|呢)",
    r"^(你.*会|你.*能)(?!.*检测|.*分析|.*测试|.*色谱|.*柱|.*产品|.*培养基|.*试剂)",
    r"(你是谁创建|谁开发|谁做的|谁教你|你的.*作者|你的.*开发者|creator|developer)",
    r"(你.*聪明|你.*笨|你.*傻|你.*厉害|你.*强大|are\s*you\s*smart)",
    r"(你.*恋爱|你.*结婚|你.*有.*男朋友|你.*有.*女朋友|你.*喜欢|你.*爱)",
    r"(你有.*感情|你有.*情绪|你有.*灵魂|你有.*意识|你有.*感觉)",
    r"(你会.*写.*代码|你会.*编程|你会.*翻译|你会.*画画|你会.*唱歌)",
    r"(你.*在哪|你.*住哪里|你.*地址|where\s*are\s*you\s*located)",
    r"(给我.*建议|帮我.*决定|推荐.*电影|推荐.*书|推荐.*音乐|有什么.*好.*推荐)",
]

def is_simple_question(question: str) -> bool:
    """检测是否是不需要知识库检索的简单提问。

    这些提问在系统提示词中已有预设的固定回复模板（如身份介绍、COA引导、价格引导、
    问候语等），无需进行 FAISS 向量检索即可直接由 LLM 基于 system prompt 回复。
    """
    q = question.strip().lower()
    if not q:
        return True
    for pattern in SIMPLE_QUESTION_PATTERNS:
        if re.search(pattern, q, re.IGNORECASE):
            return True
    return False


# ---------------- 主流程：检索 + 判定 + 生成 ----------------
async def retrieve(question: str, file_id: str) -> tuple[list[dict], str]:
    """
    返回 (citations, context_text)
    citations: [{citation_id, fileId, rank, page, snippet, score, previewUrl}]
    context_text: 供 LLM 使用的拼接上下文
    """
    vs = _load_vs(file_id)
    hits = vs.similarity_search_with_score(question, k=K)
    citations = []
    ctx_snippets = []
    scores = []
    for i, (doc, score) in enumerate(hits, start=1):
        snippet_short = (doc.page_content or "").strip()
        if len(snippet_short) > 500:
            snippet_short = snippet_short[:500] + "..."
        page = doc.metadata.get("page") or doc.metadata.get("page_number")
        citations.append({
            "citation_id": f"{file_id}-c{i}",
            "fileId": file_id,
            "rank": i,
            "page": page,
            "snippet": (doc.page_content or "")[:4000],
            "score": float(score),
            "previewUrl": f"/api/v1/pdf/page?fileId={file_id}&page={(page or 1)}&type=original",
        })
        ctx_snippets.append(f"[{i}] {snippet_short}")
        scores.append(float(score))
    context_text = "\n\n".join(ctx_snippets) if ctx_snippets else "(no hits)"

    # 规则 + LLM 复核
    ok_by_score = _score_ok(scores)
    if not ok_by_score:
        grader = _get_grader()
        grade_prompt = GRADE_PROMPT.format(context=context_text, question=question)
        decision = await grader.ainvoke([{"role": "user", "content": grade_prompt}])
        ok_by_llm = "yes" in (decision.content or "").lower()
    else:
        ok_by_llm = True

    branch = "with_context" if ok_by_llm else "no_context"
    return citations, context_text if branch == "with_context" else ""

async def answer_stream(
    question: str,
    citations: list[dict],
    context_text: str,
    branch: str,
    session_id: str | None = None
) -> AsyncGenerator[dict, None]:
    """
    以增量事件的形式产出：
      {"type":"citation", "data": {...}}
      {"type":"token", "data": "text chunk"}
      {"type":"done", "data": {"used_retrieval": bool}}
    同时：如果提供了 session_id，会把本轮问答写入内存历史。
    """
    # 先把 citations 全部发给前端（便于角标立刻出现）
    if branch == "with_context" and citations:
        for c in citations:
            yield {"type": "citation", "data": c}

    # 组装"历史 + 本轮提示"
    llm = _get_llm()
    history_msgs = get_history(session_id) if session_id else []

    if branch == "with_context" and context_text:
        user_prompt = ANSWER_WITH_CONTEXT.format(question=question, context=context_text)
    else:
        user_prompt = ANSWER_NO_CONTEXT.format(question=question)

    # 完整消息序列：system + 历史多轮 + 当前用户
    msgs = [{"role": "system", "content": SYSTEM_INSTRUCTION}]
    # 将历史逐条附加（保持 role: "user"/"assistant"）
    msgs.extend(history_msgs)
    # 当前用户问题
    msgs.append({"role": "user", "content": user_prompt})

    # 把最终生成的文本拼接出来用于写历史
    final_text_parts: list[str] = []
    # 滑动缓冲区：用于跨 delta 检测\"相关文档片段\"等占位文字
    # LLM 流式输出可能把 \"相/关/文/档/片/段\" 拆到多个 token 里
    _trailing_buffer = ""      # 保留最近输出的字符用于模式匹配
    _TRAILING_MAX = 80         # 缓冲区最大长度
    _REF_SECTION_RE = re.compile(
        r'相关文档片段|（无文本片段）|无文本片段',
        re.IGNORECASE
    )

    # 优先使用流式
    try:
        async for chunk in llm.astream(msgs):
            delta = getattr(chunk, "content", None)
            if not delta:
                continue

            # 追加到滑动缓冲区
            _trailing_buffer += delta
            if len(_trailing_buffer) > _TRAILING_MAX:
                _trailing_buffer = _trailing_buffer[-_TRAILING_MAX:]

            # 跨 delta 检测\"相关文档片段\"等占位文字
            m = _REF_SECTION_RE.search(_trailing_buffer)
            if m:
                # 找到了：截断到匹配位置之前
                cut_pos = m.start()
                # 计算这个 delta 中有多少字符是可以保留的
                delta_len = len(delta)
                overlap_in_delta = cut_pos - (len(_trailing_buffer) - delta_len)
                if overlap_in_delta > 0:
                    # 部分在当前 delta 中：只保留当前 delta 中干净的字符
                    clean_part = delta[:overlap_in_delta]
                    if clean_part:
                        final_text_parts.append(clean_part)
                        yield {"type": "token", "data": clean_part}
                elif overlap_in_delta == 0:
                    # 匹配位置正好在当前 delta 开始处：整个 delta 都是垃圾，也无需回退
                    pass
                else:
                    # 匹配位置在更早的缓冲区中（理论上不应发生），不回退已发送数据
                    pass
                break  # 立即停止流式输出

            final_text_parts.append(delta)
            yield {"type": "token", "data": delta}
    except Exception:
        # 回退：非流式整段生成
        resp = await llm.ainvoke(msgs)
        text = resp.content or ""
        # 同样需要截断非流式输出中的占位文字
        text = _REF_SECTION_RE.split(text)[0].rstrip()
        final_text_parts.append(text)
        for i in range(0, len(text), 20):
            yield {"type": "token", "data": text[i:i+20]}
            await asyncio.sleep(0.005)

    if branch == "with_context" and citations:
        imgs = []
        # 取前 2 张，避免过多（可按需改成 3）
        for c in citations[:2]:
            url = c.get("previewUrl")
            if url:
                # 生成 Markdown 图片行
                imgs.append(f"![参考页 {c.get('rank', '')}]({url})")
        if imgs:
            tail = "\n\n---\n**相关页面预览**\n\n" + "\n\n".join(imgs)
            # 作为一个额外 token 块发给前端
            yield {"type": "token", "data": tail}

    # 将本轮问答写入历史（仅在提供 session_id 时）
    if session_id:
        append_history(session_id, "user", question)
        append_history(session_id, "assistant", "".join(final_text_parts))

    yield {"type": "done", "data": {"used_retrieval": branch == "with_context"}}


async def retrieve_multi(question: str, file_ids: list[str]) -> tuple[list[dict], str]:
    """
    Retrieve from multiple knowledge bases, merging results.
    Returns (citations, context_text)
    """
    if not file_ids:
        file_ids = _list_ready_indexes()

    all_citations: list[dict] = []
    all_snippets: list[str] = []
    best_scores: list[float] = []

    for file_id in file_ids:
        try:
            vs = _load_vs(file_id)
            hits = vs.similarity_search_with_score(question, k=K)
            for i, (doc, score) in enumerate(hits, start=1):
                snippet_short = (doc.page_content or "").strip()
                if len(snippet_short) > 500:
                    snippet_short = snippet_short[:500] + "..."
                page = doc.metadata.get("page") or doc.metadata.get("page_number")
                all_citations.append({
                    "citation_id": f"{file_id}-c{i}",
                    "fileId": file_id,
                    "rank": len(all_citations) + 1,
                    "page": page,
                    "snippet": (doc.page_content or "")[:4000],
                    "score": float(score),
                    "previewUrl": f"/api/v1/pdf/page?fileId={file_id}&page={(page or 1)}&type=original",
                })
                all_snippets.append(f"[{file_id}:{i}] {snippet_short}")
                best_scores.append(float(score))
        except FileNotFoundError:
            continue

    if not all_snippets:
        return [], ""

    context_text = "\n\n".join(all_snippets)
    ok_by_score = _score_ok(best_scores)
    ok_by_llm = ok_by_score
    if not ok_by_score:
        grader = _get_grader()
        grade_prompt = GRADE_PROMPT.format(context=context_text, question=question)
        decision = await grader.ainvoke([{"role": "user", "content": grade_prompt}])
        ok_by_llm = "yes" in (decision.content or "").lower()

    if ok_by_llm:
        # Deduplicate citations, keep top K total
        return all_citations[:K], context_text
    return [], ""


# ---------------- OCR for user-uploaded files ----------------

_ocr_instance = None

def _get_ocr():
    """Lazy init PaddleOCR instance"""
    global _ocr_instance
    if _ocr_instance is None and HAS_PADDLEOCR:
        _ocr_instance = PaddleOCR(lang='ch')
    return _ocr_instance


def extract_text_from_image(file_path: str) -> str:
    """Extract text from image using PaddleOCR
    
    PaddleOCR 2.x 返回格式: [[[box, (text, confidence)], ...], ...]  # 每页一个列表
    """
    if not HAS_PADDLEOCR or not HAS_PIL:
        return ""
    ocr = _get_ocr()
    if not ocr:
        return ""
    try:
        result = ocr.ocr(file_path)
        texts = []
        if not result:
            return ""
        # result 是列表，每个元素对应一页的识别结果
        for page_result in result:
            if not page_result:
                continue
            for line in page_result:
                if not line or len(line) < 2:
                    continue
                # line 格式: [box_coordinates, (text, confidence)]
                text_info = line[1]
                if isinstance(text_info, (list, tuple)) and len(text_info) >= 2:
                    text = text_info[0]  # 第一个元素是识别文本
                    confidence = text_info[1] if len(text_info) > 1 else 0
                    # 可选：过滤低置信度结果
                    if text and confidence > 0.5:
                        texts.append(text)
                elif isinstance(text_info, dict):
                    text = text_info.get("text", "")
                    if text:
                        texts.append(text)
        return "\n".join(texts)
    except Exception as e:
        print(f"OCR image error: {e}")
        return ""


def extract_text_from_pdf_with_mineru(file_path: str) -> str:
    """使用 MinerU 从 PDF 提取文本内容"""
    try:
        # 创建临时输出目录
        import tempfile
        import subprocess
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_output_path = Path(temp_dir)
            
            # 调用 mineru 命令解析 PDF
            cmd = [
                "mineru",
                "-p", str(file_path),
                "-o", str(temp_output_path),
                "-b", "pipeline"
            ]
            
            try:
                result = subprocess.run(
                    cmd,
                    check=True,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace"
                )
                print(f"MinerU 附件解析成功: {result.stdout}")
            except subprocess.CalledProcessError as e:
                print(f"MinerU 附件解析失败: {e.stderr}")
                return ""
            
            # 查找生成的 markdown 文件
            md_files = list(temp_output_path.glob("*.md"))
            if not md_files:
                print("未找到 MinerU 生成的 Markdown 文件")
                return ""
            
            # 读取 Markdown 内容
            md_file = md_files[0]
            md_content = md_file.read_text(encoding="utf-8")
            return md_content
            
    except Exception as e:
        print(f"MinerU 附件 PDF 提取文本错误: {e}")
        return ""

def extract_text_from_pdf_file(file_path: str) -> str:
    """Extract text from PDF using MinerU"""
    return extract_text_from_pdf_with_mineru(file_path)


def extract_text_from_file(file_path: str, filename: str) -> str:
    """Extract text from uploaded file based on extension"""
    ext = Path(filename).suffix.lower()
    image_exts = {".png", ".jpg", ".jpeg", ".bmp", ".webp", ".tiff", ".tif"}
    
    if ext in image_exts:
        return extract_text_from_image(file_path)
    elif ext == ".pdf":
        return extract_text_from_pdf_file(file_path)
    elif ext == ".docx":
        # DOCX: try to extract text directly first, fallback to OCR if needed
        try:
            from docx import Document
            doc = Document(file_path)
            text = "\n".join(p.text for p in doc.paragraphs)
            if text.strip():
                return text
            # 如果直接提取为空，尝试OCR（含图片的DOCX）
            return ""
        except Exception as e:
            print(f"DOCX text extraction failed: {e}")
            return ""
    elif ext == ".xlsx":
        # XLSX: extract text from cells
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            texts = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and str(cell).strip():
                        texts.append(str(cell))
            wb.close()
            return "\n".join(texts)
        except Exception as e:
            print(f"XLSX text extraction failed: {e}")
            return ""
    elif ext in {".txt", ".md"}:
        try:
            return Path(file_path).read_text(encoding="utf-8")
        except:
            return ""
    return ""
